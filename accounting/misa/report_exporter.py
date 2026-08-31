import os
import asyncio
import logging
from django.conf import settings
from accounting.misa.browser import find_locator_in_any_frame, close_misa_popups

logger = logging.getLogger(__name__)

async def ensure_all_items_selected(page_or_frame):
    """
    Tự động kiểm tra và tích chọn 'Chọn tất cả' các vật tư/khách hàng/nhà cung cấp trong modal tham số.
    CHỈ click nút 'Chọn tất cả' / Header Checkbox, KHÔNG click dòng lẻ để tránh bỏ chọn Tất cả.
    """
    logger.info("Ensuring all items/materials 'Chọn tất cả' header checkboxes are checked...")
    return await check_all_select_all_checkboxes(page_or_frame)

async def dismiss_misa_warning_if_any(page):
    """
    Tự động phát hiện và bấm 'Đóng'/'Đồng ý' nếu MISA hiện popup cảnh báo 'Chưa chọn vật tư hàng hóa'.
    """
    try:
        warning_modal = page.locator(".ms-message-box, .dx-dialog-content, [role='dialog']:has-text('vật tư'), [role='dialog']:has-text('chưa chọn'), div:has-text('Bạn chưa chọn')").first
        if await warning_modal.is_visible(timeout=1500):
            logger.warning("DETECTED MISA WARNING POPUP: 'Chưa chọn vật tư hàng hóa' / 'Bạn chưa chọn'!")
            try:
                ss_path = os.path.join(settings.BASE_DIR, 'scratch', 'screenshots', 'error_vattuhanghoa.png')
                os.makedirs(os.path.dirname(ss_path), exist_ok=True)
                await page.screenshot(path=ss_path)
            except Exception:
                pass
            
            close_btn = warning_modal.locator("button:has-text('Đồng ý'), button:has-text('Đóng'), .ms-button-primary, span:has-text('Đóng')").first
            if await close_btn.is_visible(timeout=1500):
                await close_btn.click(force=True)
                await asyncio.sleep(1.0)
                logger.info("Dismissed MISA warning popup. Retrying item selection...")
                return True
    except Exception as e:
        logger.debug(f"No warning popup detected: {e}")
    return False

async def remove_nhat_branches(page):
    """
    Loại bỏ các tag chi nhánh có chứa ký tự '_Nhật' trong ô chọn Chi nhánh (MISA parameter form).
    (Phục hồi 100% pre-commit 773e281~1)
    """
    logger.info("Scrutinizing and removing all branch tags containing '_Nhật'...")
    total_removed = 0
    tag_xpath = "//*[contains(text(), '_Nhật')]/ancestor::*[contains(concat(' ', normalize-space(@class), ' '), ' selected-item ') or contains(concat(' ', normalize-space(@class), ' '), ' dx-tag ') or contains(concat(' ', normalize-space(@class), ' '), ' ms-tag ') or contains(concat(' ', normalize-space(@class), ' '), ' dx-tag-content ') or contains(concat(' ', normalize-space(@class), ' '), ' ms-tag-content ') or contains(concat(' ', normalize-space(@class), ' '), ' badge ')]"

    for frame in [page] + [f for f in page.frames if f != page.main_frame]:
        try:
            tag_containers = frame.locator(f"xpath={tag_xpath}")
            count = await tag_containers.count()
            if count > 0:
                logger.info(f"Found {count} branch tags containing '_Nhật' in frame.")
                for attempt in range(15):
                    tag_containers = frame.locator(f"xpath={tag_xpath}")
                    count = await tag_containers.count()
                    if count == 0:
                        break
                    clicked_any = False
                    for i in range(count):
                        tag = tag_containers.nth(i)
                        if await tag.is_visible():
                            close_btn = tag.locator("xpath=.//*[contains(@class, 'close') or contains(@class, 'remove') or contains(@class, 'clear') or contains(@class, 'dx-tag-remove-button') or contains(@class, 'mi-close') or text()='x' or text()='×']").first
                            if await close_btn.count() > 0 and await close_btn.is_visible():
                                logger.info("Removing branch tag containing '_Nhật' (click close button)")
                                await close_btn.click(force=True)
                                await asyncio.sleep(0.5)
                                clicked_any = True
                                total_removed += 1
                                break
                            else:
                                box = await tag.bounding_box()
                                if box:
                                    logger.info("Removing branch tag containing '_Nhật' (click right edge coordinates)")
                                    await page.mouse.click(box['x'] + box['width'] - 10, box['y'] + box['height'] / 2)
                                    await asyncio.sleep(0.5)
                                    clicked_any = True
                                    total_removed += 1
                                    break
                    if not clicked_any:
                        break
        except Exception as e:
            logger.warning(f"Error removing '_Nhật' branch tags in frame: {e}")

    # Sweep JS chip elements for any extra chips
    try:
        js_sweep = """
        (() => {
            let count = 0;
            const tags = Array.from(document.querySelectorAll('.ms-tag, .ms-chip, .dx-tag, [class*="tag"], [class*="chip"]'));
            tags.forEach(tag => {
                const text = (tag.textContent || '').normalize('NFC').trim();
                if ((text.includes('_Nhật') || text.includes('FUJI') || text.includes('Nippon')) && text.length < 80) {
                    const closeBtn = tag.querySelector('.ms-tag-close, .icon-close, .ms-tag-remove, .dx-tag-remove-button, [class*="close"], [class*="remove"], i, svg, span.x');
                    if (closeBtn) {
                        closeBtn.click();
                        count++;
                    } else {
                        tag.click();
                        count++;
                    }
                }
            });
            return count;
        })()
        """
        for frame in [page] + [f for f in page.frames if f != page.main_frame]:
            js_count = await frame.evaluate(js_sweep)
            if isinstance(js_count, int) and js_count > 0:
                total_removed += js_count
    except Exception:
        pass

    if total_removed > 0:
        logger.info(f"SUCCESS: Total removed {total_removed} branch tags containing '_Nhật'.")
        return True
    else:
        logger.info("Checked 'Chi nhánh' box: No branch tags containing '_Nhật' found.")
        return False

async def check_all_select_all_checkboxes(page):
    """
    Chỉ tích chọn ô checkbox nằm NGAY KẾ BÊN chữ 'Chọn tất cả' (ví dụ: 'Chọn tất cả 31355 vật tư được chọn').
    Thực hiện chậm lại 1 giây giữa các lần click theo đúng yêu cầu của người dùng.
    """
    logger.info("Targeting standalone checkboxes next to 'Chọn tất cả' text (with 1.0s delay between clicks)...")
    total_checked = 0
    js_script = """
    async () => {
        let clickedCount = 0;
        const popups = document.querySelectorAll('div.ms-popup, div.dx-popup-content, div.con-ms-popup, div.ms-dialog');
        const containers = popups.length > 0 ? popups : [document.body];
        
        for (const container of containers) {
            const textNodes = Array.from(container.querySelectorAll('*')).filter(el => {
                if (el.tagName === 'TH' || el.closest('th') || el.closest('thead')) return false;
                if (el.children.length > 3) return false;
                const txt = (el.textContent || '').normalize('NFC').trim();
                return txt.includes('Chọn tất cả');
            });
            
            for (const node of textNodes) {
                const wrapper = node.closest('.con-ms-checkbox, .ms-checkbox--content, label, div.flex, div.row, div') || node.parentElement;
                if (!wrapper || wrapper.tagName === 'TH' || wrapper.closest('th')) continue;
                
                const cb = wrapper.querySelector('span.ms-checkbox, input[type="checkbox"], div.dx-checkbox-icon') || 
                           (wrapper.parentElement ? wrapper.parentElement.querySelector('span.ms-checkbox, input[type="checkbox"], div.dx-checkbox-icon') : null);
                
                if (cb) {
                    const spanClass = cb.getAttribute('class') || '';
                    const wrapperClass = wrapper.getAttribute('class') || '';
                    const inputChecked = cb.tagName === 'INPUT' ? cb.checked : false;
                    
                    const isAlreadyChecked = spanClass.includes('checked-true') || 
                                             spanClass.includes('dx-checkbox-checked') || 
                                             wrapperClass.includes('checked-true') || 
                                             inputChecked;
                    
                    if (!isAlreadyChecked && cb.offsetWidth > 0 && cb.offsetHeight > 0) {
                        cb.click();
                        clickedCount++;
                        await new Promise(resolve => setTimeout(resolve, 1000));
                    }
                }
            }
        }
        return clickedCount;
    }
    """
    for frame in [page] + [f for f in page.frames if f != page.main_frame]:
        try:
            cnt = await frame.evaluate(js_script)
            if isinstance(cnt, int) and cnt > 0:
                total_checked += cnt
        except Exception:
            pass
    logger.info(f"Successfully checked {total_checked} standalone 'Chọn tất cả' checkboxes (with 1.0s delay).")
    return total_checked

async def select_accounts_for_so_chi_tiet(page, accounts=['111', '112', '341', '641', '642']):
    """
    Tự động chọn Bậc = 1 và chọn từng tài khoản trong báo cáo Sổ chi tiết các tài khoản (TAI_KHOAN_CT).
    Phục hồi 100% logic commit 57a0e59:
    1. Chọn Bậc = 1 qua combobox (với các fallback selector + keyboard navigation)
    2. Với TỪNG tài khoản: gõ mã tài khoản vào ô tìm kiếm 'Nhập từ khóa tìm kiếm', chờ bảng lọc, rồi click checkbox dòng khớp bằng JS.
    """
    logger.info(f"[TAI_KHOAN_CT] Setting Account Level = 1 and selecting accounts: {accounts}")

    # --- Bước 1: Chọn Bậc = 1 ---
    bac_selectors = [
        "xpath=//div[contains(@class,'ms-search-account')]//div[contains(@class,'ms-combo')]//input",
        "xpath=//span[normalize-space(text())='Bậc']/following::div[contains(@class,'ms-combo')][1]//input",
        "xpath=//th[normalize-space(text())='Bậc']/preceding::input[contains(@class,'dx-texteditor-input')][1]",
        "xpath=//div[contains(@class,'ms-combo') and not(ancestor::*[contains(@class,'ms-date')])][last()]//input",
        ".ms-search-account .ms-combo input",
        "xpath=(//div[contains(@class,'param') or contains(@class,'filter')]//input[@type='text'])[last()]",
    ]
    bac_input, _ = await find_locator_in_any_frame(page, bac_selectors, timeout=3000)
    if bac_input:
        logger.info("Clicking 'Bac' combobox to open options...")
        await bac_input.click(force=True)
        await asyncio.sleep(0.5)

        bac1_selectors = [
            "xpath=//div[contains(@class,'dx-dropdowneditor-overlay')]//div[contains(@class,'dx-item-content') and normalize-space(.)='1']",
            "xpath=//div[contains(@class,'dx-item-content') and normalize-space(text())='1']",
            "xpath=//li[contains(@class,'dx-list-item')][normalize-space(.)='1']",
            "xpath=//*[contains(@class,'ms-combo-item') or contains(@class,'ms-select-item')][normalize-space(text())='1']",
            "xpath=//div[contains(@class,'ms-dropdown')]//*[normalize-space(text())='1']"
        ]
        bac1_item, _ = await find_locator_in_any_frame(page, bac1_selectors, timeout=2000)
        if bac1_item:
            logger.info("Selecting 'Bac 1'...")
            await bac1_item.click(force=True)
        else:
            logger.warning("Could not find 'Bac 1' option. Using keyboard ArrowDown fallback...")
            try:
                await bac_input.click(force=True)
                await asyncio.sleep(0.5)
                await page.keyboard.press("Home")
                await asyncio.sleep(0.2)
                await page.keyboard.press("ArrowDown")
                await asyncio.sleep(0.2)
                await page.keyboard.press("Enter")
                logger.info("Used keyboard ArrowDown to select Bac 1 and pressed Enter.")
            except Exception as e:
                logger.error(f"Failed to select Bac 1: {str(e)}")
        await asyncio.sleep(0.5)
    else:
        logger.warning("Could not find 'Bac' combobox. Skipping Bac selection.")

    # --- Bước 2: Với TỪNG tài khoản, dùng ô tìm kiếm để lọc rồi click ---
    account_search_selectors = [
        "input[placeholder='Nhập từ khóa tìm kiếm']",
        "xpath=//input[@placeholder='Nhập từ khóa tìm kiếm']",
        "input[placeholder*='từ khóa']",
        "input[placeholder*='tìm kiếm']",
        ".ms-input-search input",
    ]

    for account_code in accounts:
        logger.info(f"[TAI_KHOAN_CT] Searching and selecting account: {account_code}")

        acc_input, _ = await find_locator_in_any_frame(page, account_search_selectors, timeout=3000)
        if not acc_input:
            logger.warning(f"Could not find account search textbox for {account_code}. Skipping.")
            continue

        try:
            await acc_input.click(force=True, click_count=3)
            await asyncio.sleep(0.1)
            await page.keyboard.press("Control+A")
            await asyncio.sleep(0.1)
            await page.keyboard.press("Backspace")
            await asyncio.sleep(0.2)
            # Dispatch JS input & change event để Vue/React reactive state xóa sạch giá trị
            await acc_input.evaluate("""el => {
                el.value = '';
                el.dispatchEvent(new Event('input', { bubbles: true }));
                el.dispatchEvent(new Event('change', { bubbles: true }));
            }""")
            await asyncio.sleep(0.3)
            await acc_input.type(account_code, delay=50)
            await asyncio.sleep(0.3)
            await page.keyboard.press("Enter")
            await asyncio.sleep(1.5)  # Chờ bảng filter hiển thị kết quả
            logger.info(f"Typed account code '{account_code}' in search box.")
        except Exception as e:
            logger.error(f"Failed to type account {account_code}: {str(e)}")
            continue

        # Click checkbox của dòng khớp chính xác bằng JS (quét toàn bộ DOM)
        clicked = False
        try:
            js_script = f"""
            (() => {{
                const rows = document.querySelectorAll('tr, [role="row"], .dx-data-row, .dx-row, .ms-tr');
                if (rows.length === 0) return false;

                for (const row of rows) {{
                    const cells = row.querySelectorAll('td, th, [role="cell"], .dx-cell, .ms-td');
                    let hasMatch = false;
                    for (const cell of cells) {{
                        if (cell.textContent.trim() === '{account_code}') {{
                            hasMatch = true;
                            break;
                        }}
                    }}
                    if (hasMatch) {{
                        const cb = row.querySelector('input[type="checkbox"], .dx-checkbox, .ms-checkbox, .checkbox, [role="checkbox"]');
                        if (cb) {{
                            let isChecked = false;
                            if (cb.tagName === 'INPUT') {{
                                isChecked = cb.checked;
                            }} else {{
                                isChecked = cb.classList.contains('dx-checkbox-checked') ||
                                            cb.classList.contains('is-checked') ||
                                            cb.classList.contains('checked') ||
                                            cb.getAttribute('aria-checked') === 'true' ||
                                            row.classList.contains('dx-selection') ||
                                            row.getAttribute('aria-selected') === 'true';
                            }}
                            if (!isChecked) {{ cb.click(); }}
                            return true;
                        }}
                        if (cells.length > 0) {{ cells[0].click(); return true; }}
                    }}
                }}

                // Nếu sau khi filter chỉ còn 1 dòng data -> click dòng đầu tiên
                let dataRow = rows[0];
                if (rows.length > 1 && (rows[0].closest('thead') || rows[0].querySelector('th'))) {{
                    dataRow = rows[1];
                }}
                const cb = dataRow.querySelector('input[type="checkbox"], .dx-checkbox, .ms-checkbox, .checkbox, [role="checkbox"]');
                if (cb) {{ cb.click(); return true; }}
                const cells2 = dataRow.querySelectorAll('td, [role="cell"], .dx-cell, .ms-td');
                if (cells2.length > 0) {{ cells2[0].click(); return true; }}
                return false;
            }})()
            """
            frames_to_check = [page] + page.frames
            for f in frames_to_check:
                try:
                    if await f.evaluate(js_script):
                        clicked = True
                        break
                except Exception:
                    pass

            if clicked:
                logger.info(f"Successfully clicked checkbox for account {account_code} via JS.")
        except Exception as e:
            logger.error(f"JS fallback for {account_code} failed: {str(e)}")

        if not clicked:
            logger.warning(f"Could not click via JS for {account_code}, trying XPath first-row fallback...")
            first_row_selectors = [
                "xpath=(//tr[contains(@class,'dx-row')]//div[contains(@class,'dx-checkbox')])[1]",
                "xpath=(//div[contains(@class,'dx-data-row')]//div[contains(@class,'dx-checkbox')])[1]",
                ".dx-checkbox:visible"
            ]
            first_row_cb, _ = await find_locator_in_any_frame(page, first_row_selectors, timeout=1500)
            if first_row_cb:
                await first_row_cb.click(force=True)
                logger.info(f"Clicked first row checkbox as fallback for account {account_code}.")
            else:
                logger.warning(f"Could not find any checkbox for {account_code}. Skipping.")

        await asyncio.sleep(0.5)

    logger.info(f"[TAI_KHOAN_CT] Finished selecting accounts: {accounts}")

async def select_account_for_tuoi_no_kh(page, account_code):
    """
    Chọn mã tài khoản đơn (ví dụ '131' hoặc '1311') trong ô Combobox 'Tài khoản *' của báo cáo TUOI_NO_KH.
    """
    logger.info(f"[TUOI_NO_KH] Selecting account '{account_code}' in 'Tài khoản *' combobox...")
    acc_selectors = [
        "xpath=//label[contains(text(), 'Tài khoản')]/ancestor::div[contains(@class, 'ms-combo')]//input",
        "xpath=//div[contains(text(), 'Tài khoản')]/ancestor::div[contains(@class, 'ms-combo')]//input",
        "xpath=//label[contains(text(), 'Tài khoản')]/following::div[contains(@class,'ms-combo')][1]//input",
        "xpath=//input[contains(@placeholder, 'Tài khoản')]",
        ".ms-combo input[placeholder*='Tài khoản']"
    ]
    acc_input, frame = await find_locator_in_any_frame(page, acc_selectors, timeout=3000)
    if not acc_input:
        logger.warning(f"[TUOI_NO_KH] Could not find 'Tài khoản *' combobox input for {account_code}.")
        return False

    try:
        await acc_input.click(force=True)
        await asyncio.sleep(0.5)
        await page.keyboard.press("Control+A")
        await page.keyboard.press("Backspace")
        await acc_input.fill("")
        await acc_input.type(account_code)
        await asyncio.sleep(1.0)

        item_selectors = [
            f"xpath=//div[contains(@class,'dx-dropdowneditor-overlay')]//*[normalize-space(text())='{account_code}']",
            f"xpath=//div[contains(@class,'dx-item-content') and normalize-space(.)='{account_code}']",
            f"xpath=//tr[contains(@class,'dx-row')]//td[normalize-space(.)='{account_code}']",
            f"xpath=//*[contains(@class,'ms-combo-item') or contains(@class,'ms-select-item')][normalize-space(text())='{account_code}']"
        ]
        item, _ = await find_locator_in_any_frame(page, item_selectors, timeout=2000)
        if item:
            await item.click(force=True)
            logger.info(f"[TUOI_NO_KH] Clicked dropdown item for account '{account_code}'.")
        else:
            logger.warning(f"[TUOI_NO_KH] Dropdown item for '{account_code}' not found directly. Pressing Enter...")
            await page.keyboard.press("Enter")
        await asyncio.sleep(0.5)
        return True
    except Exception as e:
        logger.error(f"[TUOI_NO_KH] Error selecting account '{account_code}': {str(e)}")
        return False

def merge_tuoi_no_kh_excel_files(acc_file_map, final_output_path):
    """
    Gộp các file Excel TUOI_NO_KH của từng tài khoản (131, 1311) thành 1 file duy nhất,
    đồng thời bổ sung cột 'Tài khoản' chứa mã TK tương ứng để phân biệt dữ liệu trong DB.
    """
    import openpyxl

    combined_wb = None
    combined_ws = None

    for acc_code, filepath in acc_file_map.items():
        if not os.path.exists(filepath):
            logger.warning(f"[TUOI_NO_KH Merge] Temp file {filepath} not found for account {acc_code}. Skipping.")
            continue

        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active

            header_row_idx = -1
            for r_idx in range(1, 20):
                row_vals = [str(cell.value or '').strip() for cell in ws[r_idx]]
                if 'Mã khách hàng' in row_vals:
                    header_row_idx = r_idx
                    break

            if header_row_idx == -1:
                logger.error(f"[TUOI_NO_KH Merge] Header 'Mã khách hàng' not found in {filepath}.")
                wb.close()
                continue

            if not combined_wb:
                combined_wb = openpyxl.Workbook()
                combined_ws = combined_wb.active
                combined_ws.title = ws.title

                for r in range(1, header_row_idx + 2):
                    row_vals = [cell.value for cell in ws[r]]
                    if r == header_row_idx:
                        row_vals.append('Tài khoản')
                    elif r == header_row_idx + 1:
                        row_vals.append('')
                    combined_ws.append(row_vals)

            for r in range(header_row_idx + 2, ws.max_row + 1):
                row_vals = [cell.value for cell in ws[r]]
                if any(cell is not None and str(cell).strip() != '' for cell in row_vals):
                    c0 = str(row_vals[0] or '').strip().lower()
                    if c0 in ['tổng cộng', 'tổng', 'cộng'] or c0.startswith('tổng cộng') or c0.startswith('tổng :'):
                        continue
                    row_vals.append(str(acc_code))
                    combined_ws.append(row_vals)

            wb.close()
        except Exception as e:
            logger.error(f"[TUOI_NO_KH Merge] Failed to process temp file {filepath}: {e}")

    if combined_wb:
        os.makedirs(os.path.dirname(final_output_path), exist_ok=True)
        combined_wb.save(final_output_path)
        combined_wb.close()
        logger.info(f"[TUOI_NO_KH Merge] SUCCESS: Merged accounts {list(acc_file_map.keys())} into {final_output_path}")
        return True
    return False

async def download_report_from_url(page, report_url, export_selector, output_path, prefix=None, skip_parameters=False, period_option=None, target_account=None):
    """
    Phục hồi 100% luồng thao tác Playwright pre-commit 773e281~1.
    Tích hợp luồng 5 bước tinh gọn cho Danh mục Khách hàng và Danh mục Nhân viên.
    """
    is_master_data = prefix in ['DANH_SACH_KHACH_HANG', 'DANH_SACH_NHAN_VIEN', 'KHACH_HANG', 'NHAN_VIEN']
    if is_master_data:
        skip_parameters = True
        logger.info(f"[{prefix}] Master Data export mode activated (Streamlined 5-step flow without parameter modals).")

    if prefix == 'TUOI_NO_KH' and not skip_parameters and target_account is None:
        accounts_to_fetch = getattr(settings, 'MISA_TUOI_NO_KH_ACCOUNTS', ['131', '1311'])
        if len(accounts_to_fetch) > 1:
            logger.info(f"[TUOI_NO_KH] Multi-account export mode enabled for accounts: {accounts_to_fetch}")
            acc_file_map = {}
            scratch_dir = os.path.join(settings.BASE_DIR, 'scratch', 'temp_tuoi_no')
            os.makedirs(scratch_dir, exist_ok=True)

            for acc in accounts_to_fetch:
                temp_file = os.path.join(scratch_dir, f"temp_TUOI_NO_KH_{acc}.xlsx")
                logger.info(f"[TUOI_NO_KH] Executing export for account '{acc}' -> {temp_file}")
                res = await download_report_from_url(
                    page, report_url, export_selector, temp_file,
                    prefix=prefix, skip_parameters=skip_parameters,
                    period_option=period_option, target_account=acc
                )
                if res and os.path.exists(temp_file):
                    acc_file_map[acc] = temp_file

            if acc_file_map:
                merged = merge_tuoi_no_kh_excel_files(acc_file_map, output_path)
                logger.info(f"[TUOI_NO_KH] Preserved raw account files in scratch/temp_tuoi_no/ for verification: {list(acc_file_map.values())}")
                return merged
            return False

    if report_url:
        logger.info(f"Navigating to report URL: {report_url}")
        try:
            await page.goto(report_url, timeout=30000, wait_until="load")
        except Exception as e:
            logger.warning(f"Navigation to report URL timed out or failed: {str(e)}")
        try:
            await page.wait_for_load_state("domcontentloaded", timeout=10000)
        except Exception:
            pass
            
        logger.info("Waiting 5s to check for SPA routing redirect...")
        await asyncio.sleep(5)
        
        logger.info(f"Current Page URL: {page.url}")
        try:
            title = await page.title()
            logger.info(f"Current Page Title: {title}")
        except Exception:
            pass
            
        if report_url and ("home" in page.url or "dashboard" in page.url or "verify" in page.url or "callback" in page.url or (not is_master_data and "RP" not in page.url)):
            logger.info("Auto-redirected away from target URL. Returning to report URL...")
            await close_misa_popups(page)
            await asyncio.sleep(1)
            await page.goto(report_url, timeout=30000, wait_until="load")
            await asyncio.sleep(3)
            await close_misa_popups(page)

    if "login" in page.url or "sso" in page.url or "id.misa.vn" in page.url or "amisapp.misa.vn/login" in page.url:
        logger.warning("Redirected to login page. Session might be expired.")
        return False
        
    try:
        logger.info("Waiting for initial page elements to appear...")
        await asyncio.sleep(3)
        await close_misa_popups(page)
        await asyncio.sleep(1)
        
        is_saved_report = (report_url is None)

        if is_saved_report:
            if period_option:
                logger.info(f"[{prefix}] Saved Report Mode: Opening 'Chọn tham số' to set period to '{period_option}'...")
                # Step 2: Click "Chọn tham số" button
                param_btn_selectors = [
                    "button:has-text('Chọn tham số')",
                    ".btn:has-text('Chọn tham số')",
                    "div.ms-button:has-text('Chọn tham số')",
                    "span:has-text('Chọn tham số')",
                    ".dx-button-content:has-text('Chọn tham số')",
                    "[title*='Chọn tham số']",
                    "[aria-label*='Chọn tham số']",
                    ".mi-param",
                    ".icon-feature-param"
                ]
                param_btn, frame = await find_locator_in_any_frame(page, param_btn_selectors, timeout=5000)
                if not param_btn:
                    for f in page.frames:
                        try:
                            locator = f.locator("//*[contains(text(), 'Chọn tham số')]").first
                            if await locator.is_visible(timeout=1000):
                                param_btn = locator
                                frame = f
                                break
                        except Exception:
                            continue

                if param_btn:
                    logger.info(f"Clicking 'Chon tham so' button in frame: {getattr(frame, 'name', 'main') or getattr(frame, 'url', '')}")
                    await param_btn.click(force=True)
                    await asyncio.sleep(1.5)
                else:
                    logger.warning("Could not find 'Chon tham so' button. It might already be opened.")

                # Step 3: ONLY Choose Period ("Tháng trước", "Tháng 7", etc.)
                target_period = period_option
                logger.info(f"Setting saved report period to: '{target_period}'...")
                ky_baocao_selectors = [
                    "xpath=//label[contains(text(), 'Kỳ báo cáo')]/ancestor::div[contains(@class, 'ms-combo')]//input",
                    "xpath=//div[contains(text(), 'Kỳ báo cáo')]/ancestor::div[contains(@class, 'ms-combo')]//input",
                    "xpath=//label[contains(text(), 'Kỳ')]/following::div[contains(@class,'ms-combo')][1]//input",
                    ".ms-combo input[placeholder*='Kỳ']"
                ]
                ky_input, _ = await find_locator_in_any_frame(page, ky_baocao_selectors, timeout=3000)
                if not ky_input:
                    logger.warning("Combobox 'Kỳ báo cáo' not found. Proceeding with current date parameters.")
                else:
                    await ky_input.click(force=True)
                    await asyncio.sleep(0.5)

                    target_period_vars = [target_period]
                    if "Tháng " in target_period:
                        num_part = target_period.replace("Tháng ", "").strip()
                        if num_part.isdigit():
                            num_val = int(num_part)
                            target_period_vars.append(f"Tháng {num_val}")
                            target_period_vars.append(f"Tháng {num_val:02d}")

                    period_el = None
                    selected_var = None
                    for p_var in dict.fromkeys(target_period_vars):
                        exact_period_selectors = [
                            f"xpath=//div[contains(@class,'dx-dropdowneditor-overlay') or contains(@class,'ms-combo') or contains(@class,'dx-overlay-content')]//*[contains(@class,'dx-item-content') or contains(@class,'ms-combo-item') or contains(@class,'dx-list-item-content')][normalize-space(text())='{p_var}']",
                            f"xpath=//*[contains(@class,'dx-item-content') or contains(@class,'ms-combo-item') or contains(@class,'dx-list-item-content')][normalize-space(text())='{p_var}']",
                            f"xpath=//*[contains(@class,'dx-item-content') or contains(@class,'ms-combo-item') or contains(@class,'dx-list-item-content')][contains(text(),'{p_var}')]",
                            f"text='{p_var}'"
                        ]
                        period_el, _ = await find_locator_in_any_frame(page, exact_period_selectors, timeout=1500, close_blockers=False)
                        if period_el:
                            selected_var = p_var
                            logger.info(f"Found period element matching '{p_var}' in dropdown.")
                            break

                    if period_el:
                        try:
                            await period_el.click(force=True)
                            logger.info(f"Selected period '{selected_var}' successfully via UI dropdown click.")
                        except Exception as pe:
                            logger.warning(f"Clicking period element '{selected_var}' failed: {pe}. Trying keyboard input...")
                            period_el = None

                    if not period_el:
                        logger.info(f"Dropdown click failed/not found for '{target_period}'. Trying keyboard type into combobox...")
                        try:
                            await ky_input.click(force=True, click_count=3)
                            await asyncio.sleep(0.3)
                            await ky_input.type(target_period)
                            await asyncio.sleep(0.5)
                            await page.keyboard.press("Enter")
                            await asyncio.sleep(0.5)
                            period_el = True
                            logger.info(f"Typed '{target_period}' and pressed Enter successfully into period combobox.")
                        except Exception as e:
                            logger.error(f"Failed to type '{target_period}' into period combobox: {str(e)}")

                # Step 4: Click "Đồng ý" / "Xem báo cáo" button
                view_btn_selectors = [
                    "button:has-text('Đồng ý')",
                    "button:has-text('Xem báo cáo')",
                    "div.ms-button:has-text('Đồng ý')",
                    "div.ms-button:has-text('Xem báo cáo')"
                ]
                view_btn, _ = await find_locator_in_any_frame(page, view_btn_selectors, timeout=5000)
                if view_btn:
                    logger.info("Clicking 'Dong y' / 'Xem bao cao' button...")
                    await view_btn.click(force=True)
                    await asyncio.sleep(1.5)

            # Step 5: Wait for report grid data and loading overlays to finish
            if not is_master_data:
                logger.info(f"[{prefix}] Saved Report mode: Waiting for report grid data and loading overlays to finish...")
                for loading_sel in [".dx-loadpanel", ".loading", ".ms-loading", ".dx-loadindicator", ".loading-spinner"]:
                    for f in [page] + page.frames:
                        try:
                            loc = f.locator(loading_sel).first
                            if await loc.is_visible(timeout=500):
                                await loc.wait_for(state="hidden", timeout=25000)
                        except Exception:
                            pass
                for f in [page] + page.frames:
                    try:
                        grid_loc = f.locator("tr.dx-data-row, .dx-datagrid-rowsview, .ms-grid-viewer, .grid-container, table tbody tr").first
                        if await grid_loc.is_visible(timeout=1000):
                            break
                    except Exception:
                        pass
                await asyncio.sleep(3.0)
                await close_misa_popups(page)
        else:
            # Direct URL Mode (Option 1): Full parameter configuration
            if not skip_parameters:
                # Step 2: Click "Chọn tham số" button
                param_btn_selectors = [
                    "button:has-text('Chọn tham số')",
                    ".btn:has-text('Chọn tham số')",
                    "div.ms-button:has-text('Chọn tham số')",
                    "span:has-text('Chọn tham số')",
                    ".dx-button-content:has-text('Chọn tham số')",
                    "[title*='Chọn tham số']",
                    "[aria-label*='Chọn tham số']",
                    ".mi-param",
                    ".icon-feature-param"
                ]
                param_btn, frame = await find_locator_in_any_frame(page, param_btn_selectors, timeout=5000)
                if not param_btn:
                    for f in page.frames:
                        try:
                            locator = f.locator("//*[contains(text(), 'Chọn tham số')]").first
                            if await locator.is_visible(timeout=1000):
                                param_btn = locator
                                frame = f
                                break
                        except Exception:
                            continue
                            
                os.makedirs(os.path.join(settings.BASE_DIR, 'scratch', 'screenshots'), exist_ok=True)
                ss_dir = os.path.join(settings.BASE_DIR, 'scratch', 'screenshots')

                if param_btn:
                    logger.info(f"Clicking 'Chon tham so' button in frame: {getattr(frame, 'name', 'main') or getattr(frame, 'url', '')}")
                    await param_btn.click(force=True)
                    await asyncio.sleep(1.5)
                    try:
                        await page.screenshot(path=os.path.join(ss_dir, "01_opened_parameter_modal.png"))
                    except Exception:
                        pass
                else:
                    logger.warning("Could not find 'Chon tham so' button. It might already be opened.")
                
                # Step 3: Check "Bao gồm số liệu chi nhánh phụ thuộc" checkbox (Fail-Fast for direct URL)
                logger.info("Checking 'Bao gom so lieu chi nhanh phu thuoc' checkbox status...")
                branch_checked_success = False
                try:
                    target_frame_for_checkbox = frame if frame else page
                    checkbox_label = target_frame_for_checkbox.locator("label:has-text('Bao gồm số liệu chi nhánh phụ thuộc')").first
                    if await checkbox_label.count() == 0:
                        checkbox_label = target_frame_for_checkbox.locator("label:has-text('chi nhánh phụ thuộc')").first

                    if await checkbox_label.count() > 0:
                        checkbox_span = checkbox_label.locator("span.ms-checkbox").first
                        is_already_checked = False
                        if await checkbox_span.count() > 0:
                            span_class = await checkbox_span.get_attribute("class") or ""
                            is_already_checked = "checked-true" in span_class
                            logger.info(f"Checkbox span class: '{span_class}', is_already_checked={is_already_checked}")
                        else:
                            checkbox_input = checkbox_label.locator("input[type='checkbox']").first
                            if await checkbox_input.count() > 0:
                                is_already_checked = await checkbox_input.is_checked()

                        if not is_already_checked:
                            logger.info("Checkbox not checked yet. Clicking label to toggle ON...")
                            await checkbox_label.click(force=True)
                            await asyncio.sleep(1.0)
                            if await checkbox_span.count() > 0:
                                span_class_after = await checkbox_span.get_attribute("class") or ""
                                logger.info(f"Checkbox span class after click: '{span_class_after}'")
                                is_already_checked = "checked-true" in span_class_after

                            if not is_already_checked:
                                for retry_sel in ["label:has-text('chi nhánh phụ thuộc')", ".ms-checkbox-label:has-text('chi nhánh phụ thuộc')"]:
                                    cb_retry = target_frame_for_checkbox.locator(retry_sel).first
                                    if await cb_retry.count() > 0:
                                        await cb_retry.click(force=True)
                                        await asyncio.sleep(1.0)
                                        span_class_after2 = await checkbox_span.get_attribute("class") or ""
                                        if "checked-true" in span_class_after2:
                                            branch_checked_success = True
                            else:
                                branch_checked_success = True
                        else:
                            logger.info("'Bao gom so lieu chi nhanh phu thuoc' is already checked.")
                            branch_checked_success = True
                    try:
                        await page.screenshot(path=os.path.join(ss_dir, "02_checked_branch_option.png"))
                    except Exception:
                        pass
                except Exception as chk_err:
                    logger.error(f"CRITICAL: Lỗi khi xử lý checkbox 'Bao gồm chi nhánh phụ thuộc': {chk_err}")
                    raise RuntimeError("CRITICAL: Không thể click chọn 'Bao gồm chi nhánh phụ thuộc'. Dừng tải file để tránh thiếu dữ liệu BU con.") from chk_err

                if not branch_checked_success:
                    raise RuntimeError("CRITICAL: Không thể click chọn 'Bao gồm chi nhánh phụ thuộc'. Dừng tải file để tránh thiếu dữ liệu BU con.")

                await asyncio.sleep(2.0)

                # Step 3.1: Remove branch tags containing '_Nhật'
                try:
                    await remove_nhat_branches(page)
                    try:
                        await page.screenshot(path=os.path.join(ss_dir, "03_removed_nhat_tags.png"))
                    except Exception:
                        pass
                except Exception as e:
                    logger.error(f"Error removing '_Nhật' branches: {str(e)}")

                if prefix == 'TAI_KHOAN_CT':
                    try:
                        accounts_to_select = getattr(settings, 'MISA_SO_CHI_TIET_ACCOUNTS', ['111', '112', '341', '641', '642'])
                        await select_accounts_for_so_chi_tiet(page, accounts_to_select)
                        try:
                            await page.screenshot(path=os.path.join(ss_dir, "06_account_level_and_5_accs.png"))
                        except Exception:
                            pass
                    except Exception as e:
                        logger.error(f"Error selecting accounts for TAI_KHOAN_CT: {str(e)}")

                if prefix == 'TUOI_NO_KH' and target_account:
                    try:
                        await select_account_for_tuoi_no_kh(page, target_account)
                    except Exception as e:
                        logger.error(f"Error selecting account '{target_account}' for TUOI_NO_KH: {str(e)}")

                # Step 4: Choose Period ("Tháng này", "Tháng 6", etc.) — Fail-Fast
                skip_ky_bao_cao = (prefix == 'TUOI_NO_KH')
                target_period = period_option if period_option else getattr(settings, 'MISA_REPORT_PERIOD_OPTION', 'Tháng này')
                logger.info(f"Setting report period to: '{target_period}'...")
                
                if skip_ky_bao_cao:
                    logger.info(f"[{prefix}] Skipping 'Ky bao cao' selection as per commit 57a0e59 logic.")
                else:
                    ky_baocao_selectors = [
                        "xpath=//label[contains(text(), 'Kỳ báo cáo')]/ancestor::div[contains(@class, 'ms-combo')]//input",
                        "xpath=//div[contains(text(), 'Kỳ báo cáo')]/ancestor::div[contains(@class, 'ms-combo')]//input",
                        "xpath=//label[contains(text(), 'Kỳ')]/following::div[contains(@class,'ms-combo')][1]//input",
                        ".ms-combo input[placeholder*='Kỳ']"
                    ]
                    ky_input, frame = await find_locator_in_any_frame(page, ky_baocao_selectors, timeout=3000)
                    if not ky_input:
                        raise RuntimeError("CRITICAL: Không tìm thấy combobox 'Kỳ báo cáo'. Dừng tiến trình để tránh tải sai thời gian.")

                    await ky_input.click(force=True)
                    await asyncio.sleep(0.5)

                    # Support variations of period strings (e.g. 'Tháng 6' vs 'Tháng 06')
                    target_period_vars = [target_period]
                    if "Tháng " in target_period:
                        num_part = target_period.replace("Tháng ", "").strip()
                        if num_part.isdigit():
                            num_val = int(num_part)
                            target_period_vars.append(f"Tháng {num_val}")
                            target_period_vars.append(f"Tháng {num_val:02d}")

                    period_el = None
                    selected_var = None
                    for p_var in dict.fromkeys(target_period_vars):
                        exact_period_selectors = [
                            f"xpath=//div[contains(@class,'dx-dropdowneditor-overlay') or contains(@class,'ms-combo') or contains(@class,'dx-overlay-content')]//*[contains(@class,'dx-item-content') or contains(@class,'ms-combo-item') or contains(@class,'dx-list-item-content')][normalize-space(text())='{p_var}']",
                            f"xpath=//*[contains(@class,'dx-item-content') or contains(@class,'ms-combo-item') or contains(@class,'dx-list-item-content')][normalize-space(text())='{p_var}']",
                            f"xpath=//*[contains(@class,'dx-item-content') or contains(@class,'ms-combo-item') or contains(@class,'dx-list-item-content')][contains(text(),'{p_var}')]",
                            f"text='{p_var}'"
                        ]
                        period_el, _ = await find_locator_in_any_frame(page, exact_period_selectors, timeout=1500, close_blockers=False)
                        if period_el:
                            selected_var = p_var
                            logger.info(f"Found period element matching '{p_var}' in dropdown.")
                            break

                    if period_el:
                        try:
                            await period_el.click(force=True)
                            logger.info(f"Selected period '{selected_var}' successfully via UI dropdown click.")
                        except Exception as pe:
                            logger.warning(f"Clicking period element '{selected_var}' failed: {pe}. Trying keyboard input...")
                            period_el = None

                    if not period_el:
                        logger.info(f"Dropdown click failed/not found for '{target_period}'. Trying keyboard type into combobox...")
                        try:
                            await ky_input.click(force=True, click_count=3)
                            await asyncio.sleep(0.3)
                            await ky_input.type(target_period)
                            await asyncio.sleep(0.5)
                            await page.keyboard.press("Enter")
                            await asyncio.sleep(0.5)
                            period_el = True
                            logger.info(f"Typed '{target_period}' and pressed Enter successfully into period combobox.")
                        except Exception as e:
                            logger.error(f"Failed to type '{target_period}' into period combobox: {str(e)}")
                            period_el = None

                    if not period_el:
                        raise RuntimeError(f"CRITICAL: Không tìm thấy và không thể chọn Kỳ báo cáo '{target_period}' (đã thử các biến thể {target_period_vars}). Dừng tiến trình.")

                try:
                    await page.screenshot(path=os.path.join(ss_dir, "04_selected_period.png"))
                except Exception:
                    pass

                # Step 5: Check ALL "Chọn tất cả" checkboxes and ensure items selected (skip for TAI_KHOAN_CT as per commit 57a0e59)
                if prefix != 'TAI_KHOAN_CT':
                    await ensure_all_items_selected(frame if frame else page)
                    await check_all_select_all_checkboxes(page)
                    try:
                        await page.screenshot(path=os.path.join(ss_dir, "05_checked_select_all.png"))
                    except Exception:
                        pass

                # Step 6: Click "Đồng ý" / "Xem báo cáo" button
                view_btn_selectors = [
                    "button:has-text('Đồng ý')",
                    "button:has-text('Xem báo cáo')",
                    "div.ms-button:has-text('Đồng ý')",
                    "div.ms-button:has-text('Xem báo cáo')"
                ]
                view_btn, _ = await find_locator_in_any_frame(page, view_btn_selectors, timeout=5000)
                if view_btn:
                    logger.info("Clicking 'Dong y' / 'Xem bao cao' button...")
                    await view_btn.click(force=True)
                    await asyncio.sleep(1.5)

                    # Check if MISA warning popup appears ("Chưa chọn vật tư...", "Bạn chưa chọn...")
                    if await dismiss_misa_warning_if_any(page):
                        logger.info("Re-checking all item checkboxes and re-clicking 'Đồng ý'...")
                        await ensure_all_items_selected(page)
                        await check_all_select_all_checkboxes(page)
                        view_btn_retry, _ = await find_locator_in_any_frame(page, view_btn_selectors, timeout=3000)
                        if view_btn_retry:
                            await view_btn_retry.click(force=True)
                            await asyncio.sleep(1.5)

                    logger.info("Waiting for report data to load...")
                    for loading_sel in [".dx-loadpanel", ".loading", ".ms-loading", ".dx-loadindicator", ".loading-spinner"]:
                        for f in [page] + page.frames:
                            try:
                                loc = f.locator(loading_sel).first
                                if await loc.is_visible(timeout=500):
                                    await loc.wait_for(state="hidden", timeout=25000)
                            except Exception:
                                pass
                    await asyncio.sleep(3.0)

            # Step 6.5: Select template "Mẫu chuẩn." (gear icon settings) for BAN_HANG only (Fail-Fast for direct URL)
            if prefix == 'BAN_HANG' and not skip_parameters:
                logger.info("[BAN_HANG] Selecting 'Mẫu chuẩn.' template via gear settings button...")
                gear_selectors = [
                    ".mi-setting__list-bold",
                    "div.mi-setting__list-bold",
                    "xpath=//div[contains(@class, 'mi-setting__list-bold')]",
                    "xpath=//div[contains(@class, 'mi-setting') and not(contains(@class, 'header-icon')) and not(contains(@class, 'mi-setting-2__nav'))]"
                ]
                for loading_sel in [".dx-loadpanel", ".loading", ".ms-loading"]:
                    try:
                        await page.locator(loading_sel).first.wait_for(state="hidden", timeout=10000)
                    except Exception:
                        pass

                gear_btn, _ = await find_locator_in_any_frame(page, gear_selectors, timeout=30000)
                if not gear_btn:
                    raise RuntimeError("CRITICAL: Không thể chuyển sang 'Mẫu chuẩn.'. Không tìm thấy nút bánh răng cài đặt sau 30s. File tải về sẽ bị thiếu cột. Đã dừng tiến trình.")

                logger.info("Clicking gear settings button...")
                await gear_btn.click(force=True)
                await asyncio.sleep(1.5)

                mau_chuan_item = None
                for sel in ["text=Mẫu chuẩn.", "xpath=//span[contains(text(), 'Mẫu chuẩn.')]"]:
                    loc = page.locator(sel).first
                    try:
                        if await loc.is_visible(timeout=2000):
                            mau_chuan_item = loc
                            break
                    except Exception:
                        continue

                if not mau_chuan_item:
                    mau_chuan_item = page.locator("text=Mẫu chuẩn.").first
                    try:
                        await mau_chuan_item.wait_for(state="visible", timeout=4000)
                    except Exception:
                        mau_chuan_item = None

                if not mau_chuan_item:
                    raise RuntimeError("CRITICAL: Không tìm thấy tùy chọn 'Mẫu chuẩn.' trong menu cài đặt. Đã dừng tiến trình.")

                logger.info("Selecting 'Mẫu chuẩn.' template option...")
                await mau_chuan_item.click(force=True)
                await asyncio.sleep(5.0)

        # Step 7: Open download manager to clear old history before exporting (100% Commit 57a0e59)
        logger.info("Opening download manager to clear old history...")
        download_manager_selectors = [
            "div.ms-download",
            "div.icon-feature-download",
            ".ms-download",
            ".icon-feature-download",
            ".mi-download",
            ".mi-cloud-download"
        ]
        manager_btn, manager_frame = await find_locator_in_any_frame(page, download_manager_selectors, timeout=4000)
        if manager_btn:
            await manager_btn.click(force=True)
            await asyncio.sleep(2.0)

            for clear_attempt in range(3):
                has_entries = False
                tai_tep_check = page.locator("text='Tải tệp'").first
                try:
                    has_entries = await tai_tep_check.is_visible(timeout=1500)
                except Exception:
                    pass

                if not has_entries:
                    logger.info(f"Download panel is clean (attempt {clear_attempt + 1}). No old entries to remove.")
                    break

                logger.info(f"Found old entries in download panel (attempt {clear_attempt + 1}). Clearing...")
                clear_btn_selectors = [
                    "text='Xóa hết lịch sử tải tệp'",
                    ".clear-all",
                    ".clear-all--text",
                    "div:has-text('Xóa hết lịch sử tải tệp')"
                ]
                clear_btn, clear_frame = await find_locator_in_any_frame(page, clear_btn_selectors, timeout=2000)
                if clear_btn:
                    logger.info("Clicking 'Xóa hết lịch sử tải tệp' to clear old downloads...")
                    await clear_btn.click(force=True)
                    await asyncio.sleep(1.5)

                    confirm_btn_selectors = [
                        "button:has-text('Có')",
                        ".ms-button:has-text('Có')",
                        ".dx-button-content:has-text('Có')",
                        "text='Có'",
                        "span:has-text('Có')"
                    ]
                    confirm_btn, confirm_frame = await find_locator_in_any_frame(page, confirm_btn_selectors, timeout=3000)
                    if confirm_btn:
                        logger.info("Clicking 'Có' on deletion confirmation popup...")
                        await confirm_btn.click(force=True)
                        await asyncio.sleep(2.0)
                    else:
                        await asyncio.sleep(1.5)
                else:
                    await asyncio.sleep(1.0)
                    break

            logger.info("Closing download manager panel...")
            await manager_btn.click(force=True)
            await asyncio.sleep(1.0)

        # Step 8: Click Excel icon dropdown button (100% Commit 57a0e59)
        if is_master_data:
            logger.info(f"[{prefix}] Looking for Master Data export button (.mi-s1-file-export / .ls-file-export)...")
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            downloads = []
            page.on("download", lambda dl: downloads.append(dl))

            created_pages = []
            page.context.on("page", lambda p: (created_pages.append(p), p.on("download", lambda dl: downloads.append(dl))))

            clicked = False
            for f in [page] + page.frames:
                try:
                    res = await f.evaluate("""() => {
                        const el = document.querySelector('.mi-s1-file-export, .ls-file-export, [class*="file-export"]');
                        if (el) {
                            const target = el.closest('.ls-toolbar-icon-btn') || el;
                            target.click();
                            return true;
                        }
                        return false;
                    }""")
                    if res:
                        logger.info(f"[{prefix}] Successfully clicked Master Data export icon in frame: {getattr(f, 'url', '') or getattr(f, 'name', '')}")
                        clicked = True
                        break
                except Exception:
                    continue

            if not clicked:
                master_btn_selectors = [
                    ".mi-s1-file-export",
                    ".ls-file-export",
                    ".ls-toolbar-icon-btn:has(.mi-s1-file-export)",
                    "div.ls-file-export",
                    "div[class*='file-export']",
                    ".mi-s1.mi-s1-file-export",
                    ".ls-toolbar-icon-btn"
                ]
                export_btn, frame = await find_locator_in_any_frame(page, master_btn_selectors, timeout=5000, close_blockers=False)
                if export_btn:
                    await export_btn.click(force=True)
                    clicked = True

            if not clicked:
                logger.error(f"[{prefix}] Could not find or click Master Data export icon.")
                return False

            logger.info(f"[{prefix}] Waiting up to 35 seconds for MISA to prepare data and trigger download...")
            download_saved = False
            for sec in range(35):
                await asyncio.sleep(1)
                if downloads:
                    dl = downloads[0]
                    await dl.save_as(output_path)
                    logger.info(f"SUCCESS: Successfully downloaded Master Data report to: {output_path} (Size: {os.path.getsize(output_path)} bytes)")
                    download_saved = True
                    break

            for p in created_pages:
                try:
                    await p.close()
                except Exception:
                    pass

            if download_saved:
                return True
            else:
                logger.error(f"[{prefix}] Master Data download event did not fire within 35 seconds.")
                return False
        else:
            logger.info("Looking for Excel export icon dropdown button...")
            excel_btn_selectors = [
                "[title*='Xuất khẩu']",
                "[title*='Xuất Excel']",
                "button:has-text('Xuất khẩu')",
                "div.ms-dropdown:has-text('Xuất khẩu')",
                ".mi-export__excel-bold",
                ".mi-excel",
                ".mi-icon-excel",
                ".mi-export",
                ".icon-excel",
                ".btn-excel"
            ]
            if export_selector:
                excel_btn_selectors.insert(0, export_selector)

            excel_btn, frame = await find_locator_in_any_frame(page, excel_btn_selectors, timeout=10000, close_blockers=False)
            if not excel_btn:
                for f in page.frames:
                    try:
                        locator = f.locator("xpath=//*[contains(@class,'mi-excel') or contains(@class,'export') or contains(text(),'Xuất khẩu')]").first
                        if await locator.is_visible(timeout=2000):
                            excel_btn = locator
                            frame = f
                            break
                    except Exception:
                        continue

            if not excel_btn:
                logger.error("Could not find Excel icon dropdown button in any frame.")
                return False

            logger.info(f"Clicking Excel icon button in frame: {getattr(frame, 'name', 'main') or getattr(frame, 'url', '')}")
            await excel_btn.click(force=True)
            await asyncio.sleep(2.5)

            # Step 8: Click Excel export option (dạng báo cáo / dạng tổng hợp cho TUOI_NO_KH, dạng dữ liệu cho các báo cáo khác)
            if prefix == 'TUOI_NO_KH':
                dropdown_selectors = [
                    "text='Xuất Excel (dạng báo cáo)'",
                    "span:has-text('Xuất Excel (dạng báo cáo)')",
                    "div:has-text('Xuất Excel (dạng báo cáo)')",
                    ".dx-menu-item-text:has-text('Xuất Excel (dạng báo cáo)')",
                    "text='Mẫu tổng hợp'",
                    "span:has-text('Mẫu tổng hợp')",
                    "text='Excel'",
                    "span:has-text('Excel')",
                    "div:has-text('Excel')"
                ]
            else:
                dropdown_selectors = [
                    "text='Xuất Excel (dạng dữ liệu)'",
                    "span:has-text('Xuất Excel (dạng dữ liệu)')",
                    "div:has-text('Xuất Excel (dạng dữ liệu)')",
                    ".dx-menu-item-text:has-text('Xuất Excel (dạng dữ liệu)')",
                    "xpath=//*[contains(text(), 'dạng dữ liệu')]",
                    "text='Excel'",
                    "span:has-text('Excel')",
                    "div:has-text('Excel')"
                ]
            dropdown_item = None
            for sel in dropdown_selectors:
                loc, _ = await find_locator_in_any_frame(page, [sel], timeout=1500, close_blockers=False)
                if loc:
                    try:
                        txt = (await loc.inner_text()).strip()
                        if 'nhập' in txt.lower():
                            logger.info(f"Skipping import item: '{txt}'")
                            continue
                        dropdown_item = loc
                        break
                    except Exception:
                        dropdown_item = loc
                        break

            if dropdown_item:
                item_text = await dropdown_item.inner_text()
                logger.info(f"Clicking Excel export item '{item_text.strip()}' for prefix '{prefix}'...")
                await dropdown_item.click(force=True)
                await asyncio.sleep(2.0)

            # Check for options dialog "Đồng ý" / "Xuất khẩu" button
            for _ in range(5):
                dialog_clicked = False
                try:
                    for f in [page] + page.frames:
                        res = await f.evaluate("""() => {
                            const pop = document.querySelector('.popup-application-ui, .con-ms-popup.popup-is-show, .ms-popup');
                            if (pop && pop.offsetWidth > 0 && pop.offsetHeight > 0) {
                                const buttons = Array.from(pop.querySelectorAll('button, div.ms-button, a, span'));
                                const okBtn = buttons.find(b => {
                                    const t = (b.textContent || '').trim().normalize('NFC');
                                    return t === 'Đồng ý' || t === 'Xuất khẩu';
                                });
                                if (okBtn) {
                                    okBtn.click();
                                    return true;
                                }
                            }
                            return false;
                        }""")
                        if res:
                            dialog_clicked = True
                            logger.info("Found and clicked confirmation button in export options dialog via DOM evaluation.")
                            break
                except Exception:
                    pass

                if not dialog_clicked:
                    agree_btn_selectors = [
                        "button:has-text('Đồng ý')",
                        "button:has-text('Xuất khẩu')",
                        ".btn:has-text('Đồng ý')",
                        ".ms-button:has-text('Đồng ý')",
                        ".ms-button:has-text('Xuất khẩu')",
                        ".dx-button-content:has-text('Đồng ý')",
                        ".dx-button-content:has-text('Xuất khẩu')",
                        "span:has-text('Đồng ý')",
                        "span:has-text('Xuất khẩu')"
                    ]
                    agree_btn, _ = await find_locator_in_any_frame(page, agree_btn_selectors, timeout=1500, close_blockers=False)
                    if agree_btn:
                        logger.info("Found confirmation button in options dialog via locator. Clicking it...")
                        await agree_btn.click(force=True)
                        dialog_clicked = True
                
                if dialog_clicked:
                    await asyncio.sleep(1.5)
                else:
                    break

        # Step 9: Open download manager panel if not open and click "Tải tệp"
        wait_seconds = 15
        logger.info(f"Waiting {wait_seconds} seconds for MISA to start generating file in background...")
        await asyncio.sleep(wait_seconds)

        download_manager_selectors = [
            ".ms-download.header-menu-icon",
            ".icon-feature-download",
            ".ms-download",
            "xpath=//div[contains(@class,'header-menu-icon') and contains(@class,'ms-download')]",
            "[title*='Tải tệp']",
            "[aria-label*='Tải tệp']"
        ]

        panel_indicators = [
            "text='Tải tệp Excel, tệp in,...'",
            "text='Xóa hết lịch sử tải tệp'",
            "text='Đang tạo đường dẫn tải tệp...'",
            ".con-ms-download:visible",
            ".ms-download-list:visible"
        ]

        # Kiểm tra và mở panel nếu chưa mở
        panel_loc, _ = await find_locator_in_any_frame(page, panel_indicators, timeout=800, close_blockers=False)
        if not panel_loc:
            logger.info("Opening download manager panel...")
            mgr_btn, _ = await find_locator_in_any_frame(page, download_manager_selectors, timeout=2500, close_blockers=False)
            if mgr_btn:
                await mgr_btn.click(force=True)
                await asyncio.sleep(1.5)

        logger.info("Waiting for file generation to complete and capturing download...")
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        saved = False

        download_event = asyncio.Event()
        captured_downloads = []

        def on_dl(dl):
            logger.info(f"Captured download event: {dl.suggested_filename}")
            captured_downloads.append(dl)
            download_event.set()

        def on_aux_page(p):
            p.on("download", on_dl)

        page.on("download", on_dl)
        if hasattr(page, 'context') and page.context:
            page.context.on("page", on_aux_page)

        download_btn_selectors = [
            "xpath=(//a[normalize-space(text())='Tải tệp' or starts-with(normalize-space(text()),'Tải tệp')])[1]",
            "xpath=(//button[normalize-space(text())='Tải tệp' or starts-with(normalize-space(text()),'Tải tệp')])[1]",
            "xpath=(//span[normalize-space(text())='Tải tệp' or starts-with(normalize-space(text()),'Tải tệp')])[1]",
            "xpath=(//div[normalize-space(text())='Tải tệp'])[1]",
            "text='Tải tệp'",
            "text='Tải về'",
        ]

        for attempt in range(60): # Chờ tối đa 90 giây
            # 1. Đảm bảo panel luôn mở (kiểm tra mỗi 5 attempt hoặc khi không thấy indicator)
            panel_loc, _ = await find_locator_in_any_frame(page, ["text='Tải tệp Excel, tệp in,...'", "text='Xóa hết lịch sử tải tệp'"], timeout=300, close_blockers=False)
            if not panel_loc or (attempt > 0 and attempt % 10 == 0):
                mgr_btn, _ = await find_locator_in_any_frame(page, download_manager_selectors, timeout=1000, close_blockers=False)
                if mgr_btn:
                    await mgr_btn.click(force=True)
                    await asyncio.sleep(1.0)

            # 2. Tìm nút Tải tệp sẵn sàng
            btn_loc, frame = await find_locator_in_any_frame(page, download_btn_selectors, timeout=800, close_blockers=False)
            if btn_loc:
                try:
                    txt = (await btn_loc.text_content() or '').strip()
                    txt_lower = txt.lower()
                    if not any(k in txt_lower for k in ['đang tạo', 'đang xử lý', 'xóa hết', 'tải tệp excel', 'ava', 'trợ lý']):
                        logger.info(f"Found ready 'Tải tệp' button (text: '{txt}') at attempt {attempt + 1}. Clicking...")
                        
                        # Thử lấy href trực tiếp
                        try:
                            href = await btn_loc.get_attribute('href')
                            if not href:
                                href = await btn_loc.evaluate("el => el.closest('a') ? el.closest('a').href : ''")
                            if href and href.startswith('http') and not href.endswith('#'):
                                logger.info(f"Direct URL found: {href[:80]}... Fetching with APIRequestContext...")
                                resp = await page.request.get(href)
                                if resp.status == 200:
                                    with open(output_path, 'wb') as f:
                                        f.write(await resp.body())
                                    if os.path.exists(output_path) and os.path.getsize(output_path) > 100:
                                        saved = True
                                        logger.info(f"SUCCESS: Saved file via direct URL: {output_path} ({os.path.getsize(output_path)} bytes)")
                                        break
                        except Exception as he:
                            logger.debug(f"Direct request failed: {he}")

                        # Click DOM và Playwright đồng thời
                        try:
                            await btn_loc.evaluate("el => (el.closest('a') || el).click()")
                        except Exception:
                            pass
                        try:
                            await btn_loc.click(force=True)
                        except Exception:
                            pass

                        try:
                            await asyncio.wait_for(download_event.wait(), timeout=15.0)
                            if captured_downloads:
                                dl = captured_downloads[0]
                                await dl.save_as(output_path)
                                if os.path.exists(output_path) and os.path.getsize(output_path) > 100:
                                    saved = True
                                    logger.info(f"SUCCESS: Saved download via event listener: {output_path} ({os.path.getsize(output_path)} bytes)")
                                    break
                        except asyncio.TimeoutError:
                            logger.warning("Download event not fired after click. Continuing loop...")
                except Exception as eval_err:
                    logger.debug(f"Button evaluation error: {eval_err}")
            else:
                if (attempt + 1) % 5 == 0:
                    logger.info(f"Attempt {attempt + 1}: Waiting for MISA file generation...")

            if saved:
                break
            await asyncio.sleep(1.5)

        try:
            page.remove_listener("download", on_dl)
            if hasattr(page, 'context') and page.context:
                page.context.remove_listener("page", on_aux_page)
        except Exception:
            pass

        if not saved:
            logger.error("Could not download file via direct URL or download event.")
            return False

        logger.info(f"SUCCESS: Successfully downloaded and saved report to: {output_path}")

        # Close download manager panel to clean UI
        try:
            manager_btn, _ = await find_locator_in_any_frame(page, download_manager_selectors, timeout=2000)
            if manager_btn:
                logger.info("Closing download manager panel...")
                await manager_btn.click(force=True)
                await asyncio.sleep(1.0)
        except Exception:
            pass

        return True

    except Exception as e:
        logger.error(f"Error during report download: {str(e)}")
        return False
